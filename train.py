# coding=gbk
from __future__ import print_function
import argparse
import torch
import torch.nn as nn
import torch.optim as optim
from torch.autograd import Variable
from multidataloader import train_loader, val_loader
from EfficientPBAN import Net
import os
import openpyxl
import pandas as pd
import numpy as np
from scipy.stats import spearmanr, pearsonr, kendalltau
from sklearn.metrics import mean_squared_error
from torch.cuda.amp import autocast, GradScaler
scaler = GradScaler()

# ==================== 参数解析 ====================
def get_args():
    parser = argparse.ArgumentParser(description='PyTorch Training')
    parser.add_argument('--batch-size', type=int, default=8)
    parser.add_argument('--test-batch-size', type=int, default=1)
    parser.add_argument('--epochs', type=int, default=30)
    parser.add_argument('--lr', type=float, default=1e-4)  # 接着第20 epoch继续训练
    parser.add_argument('--momentum', type=float, default=0.9)
    parser.add_argument('--no-cuda', action='store_true', default=False)
    parser.add_argument('--seed', type=int, default=0)
    parser.add_argument('--log-interval', type=int, default=10)
    parser.add_argument('--resume', action='store_true', default=True, help='resume from checkpoint')
    return parser.parse_args()

# ==================== 训练函数 ====================
def train(epoch):
    mymodel.train()
    for batch_idx, (data1, data2, target) in enumerate(train_loader):
        if args.cuda:
            data1, data2, target = data1.cuda(), data2.cuda(), target.cuda()
        data1, data2, target = Variable(data1), Variable(data2), Variable(target)

        optimizer.zero_grad()
        with autocast():
            output = mymodel(data1, data2)
            #output = mymodel(data1)
            loss = mse_loss(output.float(), target.float())

        scaler.scale(loss).backward()
        scaler.step(optimizer)
        scaler.update()

        if batch_idx % args.log_interval == 0:
            print(f'Train Epoch: {epoch} [{batch_idx * len(data1)}/{len(train_loader.dataset)} '
                  f'({100. * batch_idx / len(train_loader):.0f}%)]\tLoss: {loss.item():.6f}')
    return loss.item()

# ==================== 验证函数 ====================
def val():
    mymodel.eval()
    test_loss = 0
    int_ = 0
    row = 0
    outputs_all = []
    targets_all = []

    for data1, data2, target in val_loader:
        if args.cuda:
            data1, data2, target = data1.cuda(), data2.cuda(), target.cuda()
        data1, data2, target = Variable(data1), Variable(data2), Variable(target)

        output = mymodel(data1, data2)
        #output = mymodel(data1)
        int_ += 1
        test_loss += mse_loss(output.float(), target.float()).item()

        output_np = output.detach().cpu().numpy()
        target_np = target.detach().cpu().numpy()

        outputs_all.extend(output_np)
        targets_all.extend(target_np)

        for i in range(len(output)):
            sheet0.cell(i + row + 1, 2).value = target[i].item()
            sheet0.cell(i + row + 1, 1).value = output[i].item()
        row += len(output)

    workbook.save(excel_path)

    outputs_all = np.array(outputs_all).flatten()
    targets_all = np.array(targets_all).flatten()
    srcc, _ = spearmanr(outputs_all, targets_all)
    plcc, _ = pearsonr(outputs_all, targets_all)
    krcc, _ = kendalltau(outputs_all, targets_all)
    rmse = np.sqrt(mean_squared_error(targets_all, outputs_all))

    test_loss /= int_
    print(f'\nVal set: Average loss: {test_loss:.8f}')
    print(f'SRCC: {srcc:.4f}, PLCC: {plcc:.4f}, KRCC: {krcc:.4f}, RMSE: {rmse:.4f}\n')
    return test_loss,srcc

# ==================== 主程序入口 ====================
if __name__ == "__main__":
    args = get_args()
    args.cuda = not args.no_cuda and torch.cuda.is_available()

    torch.manual_seed(args.seed)
    if args.cuda:
        torch.cuda.manual_seed(args.seed)

    device = torch.device('cuda:0' if args.cuda else 'cpu')
    print("Using device:", device)

    result_dir = 'C:/Users/18810/Desktop/PBAN/nriqa/results'
    os.makedirs(result_dir, exist_ok=True)
    csv_path = os.path.join(result_dir, 'FRnewres50_123_RealSRQBT.csv')
    excel_path = os.path.join(result_dir, 'FRnewres50_123_RealSRQBT.xlsx')
    model_save_path = os.path.join(result_dir, 'FRnewres50_123_RealSRQBT.pth')

    df = pd.DataFrame(columns=['step', 'train Loss', 'test Loss'])
    if not os.path.exists(csv_path):
        df.to_csv(csv_path, index=False)

    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)
    sheet0.column_dimensions['A'].width = 15

    mymodel = Net().to(device)
    optimizer = torch.optim.Adam(mymodel.parameters(), lr=args.lr, weight_decay=1e-4)
    scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=5, gamma=0.5)
    mse_loss = torch.nn.MSELoss()

    start_epoch = 1
    best = float('inf')

    # ======== 断点续训 / 从头 ========
    if args.resume and os.path.exists(model_save_path):
        print(f"==> Loading checkpoint from {model_save_path}")
        checkpoint = torch.load(model_save_path, map_location=device)

        if isinstance(checkpoint, dict) and 'model_state_dict' in checkpoint:
            # 完整 checkpoint
            mymodel.load_state_dict(checkpoint['model_state_dict'])
            optimizer.load_state_dict(checkpoint['optimizer_state_dict'])
            scheduler.load_state_dict(checkpoint['scheduler_state_dict'])
            start_epoch = checkpoint['epoch'] + 1
            best = checkpoint.get('best_loss', float('inf'))
            print(f"Resuming from epoch {start_epoch}, best loss {best:.6f}")
        else:
            # 只有 state_dict
            mymodel.load_state_dict(checkpoint)
            print("Loaded model weights only (no optimizer/scheduler state). Starting fresh optimizer.")

    for epoch in range(start_epoch, args.epochs + 1):
        step = f"Step[{epoch}]"
        t1_loss = train(epoch)
        t2_loss, val_srcc = val()

        pd.DataFrame([[step, f"{t1_loss:.8f}", f"{t2_loss:.8f}"]]).to_csv(csv_path, mode='a', header=False, index=False)

        if val_srcc > best:
            best = val_srcc
            torch.save({
                'epoch': epoch,
                'model_state_dict': mymodel.state_dict(),
                'optimizer_state_dict': optimizer.state_dict(),
                'scheduler_state_dict': scheduler.state_dict(),
                'best_loss': best
            }, model_save_path)
            print(f"Saved new best model with loss {best:.6f}")

        scheduler.step()
